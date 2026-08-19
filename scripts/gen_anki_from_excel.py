#!/usr/bin/env python3
"""
通用 Anki 题库生成器（基于猪猪模板 V10，Excel 数据源）
======================================================

用法：
    pip install openpyxl genanki
    python gen_anki_from_excel.py 你的题库.xlsx

功能：
    - 读取 Excel 题库，自动按题型分为 单选题/多选题/判断题
    - 选项打乱 + 重标 A/B/C/D + 答案字母同步更新
    - 保命题排前面（红色徽章）
    - 判断题显示「错误说法」和「正确说法」
    - 生成 3 个 apkg 文件（单选题/多选题/判断题）

Excel 格式要求（列名必须完全一致）：
    | 列名      | 说明                                          |
    |----------|-----------------------------------------------|
    | 题型     | "单选题" / "多选题" / "判断题"                  |
    | 题干     | 题目文本                                       |
    | 选项     | 单选/多选选项，换行分隔，如 A.xxx\\nB.xxx        |
    | 答案     | 正确答案字母，如 "D" 或 "ABD" 或 "正确"          |
    | 题目依据  | 解析文本（可选）                                |
    | 备注     | 含"保命题"则标记为保命题（可选）                  |
    | 判断题解析 | 判断题的正确说法文本（可选）                     |

    其他列（如 二级纲要、题目分类、说明 等）也会被自动识别。
"""

import os
import random
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from anki_template_v10 import (
    create_model, build_options_html, build_judge_options_html,
    build_remark_html, build_baoming_html, build_wrong_point_html,
    build_right_point_html, convert_answer, esc,
)
import genanki
import openpyxl

random.seed(42)

BAOMING_KEYWORDS = ["保命", "红线", "保命法则", "保命规则", "十条红线", "十项保命"]


def is_baoming(q):
    for kw in BAOMING_KEYWORDS:
        if (kw in str(q.get("二级纲要", ""))
                or kw in str(q.get("题目分类", ""))
                or kw in str(q.get("题干", ""))
                or kw in str(q.get("备注", ""))):
            return True
    return False


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["总表"] if "总表" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("题干"):
            questions.append(d)
    print(f"📖 读取 {len(questions)} 道题 from {os.path.basename(filepath)}")
    return questions


def build_decks(questions, deck_prefix="题库"):
    out_dir = os.path.join(SCRIPT_DIR, "output")
    os.makedirs(out_dir, exist_ok=True)
    model = create_model()

    single_notes, multi_notes, judge_notes = [], [], []

    for q in questions:
        qtype = q.get("题型", "")
        stem = q.get("题干", "")
        answer = q.get("答案", "")
        analysis = q.get("题目依据", "") or q.get("说明", "")
        bm = is_baoming(q)
        judge_analysis = q.get("判断题解析", "")
        baoming_html = build_baoming_html(bm)

        if qtype == "单选题":
            opts_html, letter_map = build_options_html(q.get("选项", ""), is_multi=False)
            remark_html = build_remark_html(analysis)
            updated_answer = "".join(letter_map.get(c, c) for c in convert_answer(answer))
            note = genanki.Note(
                model=model,
                fields=[esc(stem), opts_html, updated_answer, esc(analysis), baoming_html, remark_html, "", "", ""],
                guid=genanki.guid_for("v10_single", stem, answer),
            )
            single_notes.append((note, bm))

        elif qtype == "多选题":
            opts_html, letter_map = build_options_html(q.get("选项", ""), is_multi=True)
            remark_html = build_remark_html(analysis)
            updated_answer = "".join(letter_map.get(c, c) for c in convert_answer(answer))
            note = genanki.Note(
                model=model,
                fields=[esc(stem), opts_html, updated_answer, esc(analysis), baoming_html, remark_html, "", "", ""],
                guid=genanki.guid_for("v10_multi", stem, answer),
            )
            multi_notes.append((note, bm))

        elif qtype == "判断题":
            opts_html = build_judge_options_html()
            remark_html = build_remark_html(analysis)
            wrong_html = build_wrong_point_html(stem, judge_analysis)
            right_html = build_right_point_html(judge_analysis)
            note = genanki.Note(
                model=model,
                fields=[esc(stem), opts_html, convert_answer(answer), esc(analysis), baoming_html,
                        remark_html, wrong_html, right_html, ""],
                guid=genanki.guid_for("v10_judge", stem, answer),
            )
            judge_notes.append((note, bm))

    for notes in [single_notes, multi_notes, judge_notes]:
        notes.sort(key=lambda item: 0 if item[1] else 1)

    decks_info = {
        "单选题": (1610002010, f"{deck_prefix}-单选题", single_notes),
        "多选题": (1610002020, f"{deck_prefix}-多选题", multi_notes),
        "判断题": (1610002030, f"{deck_prefix}-判断题", judge_notes),
    }

    for name, (did, dname, notes_tuples) in decks_info.items():
        if not notes_tuples:
            print(f"⏭️  {dname}: 无题目，跳过")
            continue
        deck = genanki.Deck(did, dname)
        for n, _ in notes_tuples:
            deck.add_note(n)
        pkg = genanki.Package(deck)
        path = os.path.join(out_dir, f"{dname}.apkg")
        pkg.write_to_file(path)
        print(f"✅ {dname}: {len(notes_tuples)}题, {os.path.getsize(path) / 1024:.1f} KB")

    total = len(single_notes) + len(multi_notes) + len(judge_notes)
    print(f"\n📊 总计: {total} 题")
    print(f"📁 输出目录: {out_dir}")
    return out_dir


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("❌ 请提供 Excel 文件路径")
        print("   用法: python gen_anki_from_excel.py 你的题库.xlsx")
        sys.exit(1)
    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"❌ 文件不存在: {filepath}")
        sys.exit(1)
    basename = os.path.splitext(os.path.basename(filepath))[0]
    deck_prefix = basename.replace("合并", "题库").replace("题库", "题库")
    if "题库" not in deck_prefix:
        deck_prefix = deck_prefix + "题库"
    questions = read_excel(filepath)
    build_decks(questions, deck_prefix=deck_prefix)


if __name__ == "__main__":
    main()
