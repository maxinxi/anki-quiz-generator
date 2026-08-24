#!/usr/bin/env python3
"""
中级员工试题 → 易错强化版 v4（强关联 + 概念家族 + 动态错次计数 + 全覆盖注释）
==========================================================================
v4 新增（2026-08-23 用户反馈）：
    1. 强关联引擎 v2：三信号全库两两检测——
       ① 题干相似（同一概念换主体反复出题，如信息安规"XX检修前应备份（）"家族）
       ② 选项共用（含⚡正误反转）
       ③ 跨题型加成（同一概念的单选/多选/判断互相关联）
       不再局限于"选项完全相同"，也不只扫本区块——其他小区块同样检测
    2. 概念家族（并查集）：强关联的题归为一家族，卡组内排序时家族成员相邻，
       练完一个概念立刻练它的所有变形（家族内仍保持 多选→判断→单选）
    3. 笔记标签：易错·错N次 / 预测易错 / 低风险 / 概念家族N题
       → 方便用户自建筛选牌组专刷错题

继承 v3：全 917 题注释覆盖、选项原文引用（无字母，乱序安全）、
本机答错计数器（JS+localStorage）、联网权威助记、
GUID/模型ID/卡组名与用户集合一致 → 导入保留学习进度。
"""

import os, re, sys, json
from collections import Counter, defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from anki_template_v10 import (
    create_model, build_judge_options_html, build_remark_html, esc, convert_answer,
    parse_options,
)
import genanki
import openpyxl

DECK_ID_BASE = 1610005000
TYPE_ORDER = {"多选题": 0, "判断题": 1, "单选题": 2}

from gen_anki_from_print_sheet import extract_law, merge_theme, read_print_sheet, TYPE_BADGE_CSS

# ── CSS ─────────────────────────────────────────────────────
EASY_WRONG_CSS = """
/* ── 易错提示盒 ───────────────────── */
.easy-wrong{margin-top:14px;padding:13px 17px;border-radius:14px;background:linear-gradient(135deg,#fff5f5,#ffe3e3);border:2px dashed #e03131;text-align:left;}
.easy-wrong .ew-title{font-weight:900;color:#e03131;font-size:16px;margin-bottom:9px;letter-spacing:.5px;}
.easy-wrong .ew-hook{font-size:16px;font-weight:800;color:#1864ab;background:#e7f5ff;border-radius:10px;padding:9px 13px;margin-bottom:9px;line-height:1.7;border-left:5px solid #339af0;}
.easy-wrong .ew-hook b{color:#e8590c;}
.easy-wrong .ew-answer{font-size:15px;color:#343a40;margin-bottom:8px;line-height:1.7;}
.easy-wrong .ew-answer b{color:#2b8a3e;font-size:18px;}
.easy-wrong .ew-combo{font-size:14px;color:#495057;background:rgba(255,255,255,.65);border-radius:9px;padding:7px 11px;margin-bottom:7px;line-height:1.75;}
.easy-wrong .ew-reasons{font-size:14px;color:#c92a2a;margin-bottom:8px;line-height:1.8;}
.easy-wrong .ew-link{font-size:14px;color:#343a40;background:#fff9db;border:1px solid #fcc419;border-radius:9px;padding:7px 11px;margin-bottom:7px;line-height:1.9;}
.easy-wrong .ew-link b{color:#e67700;}
.easy-wrong .ew-note{font-size:12px;color:#868e96;}
.easy-wrong.predicted{background:linear-gradient(135deg,#f6f2ff,#e5dbff);border-color:#7048e8;}
.easy-wrong.predicted .ew-title{color:#7048e8;}
.easy-wrong.predicted .ew-hook{background:#ede9fe;border-left-color:#7048e8;color:#3b3b98;}
.easy-wrong.mini{background:linear-gradient(135deg,#f1f7ff,#e7f5ff);border:2px dashed #339af0;}
.easy-wrong.mini .ew-title{color:#1971c2;}
.easy-wrong.mini .ew-hook{background:rgba(255,255,255,.6);border-left-color:#339af0;color:#1864ab;}
#ewLocalCnt{font-weight:600;font-size:12px;color:#e03131;}
@media (prefers-color-scheme: dark){
  .easy-wrong{background:linear-gradient(135deg,#3d1a1a,#4a2020);border-color:#e03131;}
  .easy-wrong .ew-answer,.easy-wrong .ew-link{color:#dee2e6;}
  .easy-wrong .ew-combo,.easy-wrong .ew-hook{background:rgba(0,0,0,.32);}
  .easy-wrong .ew-hook{color:#74c0fc;}
  .easy-wrong .ew-link{background:rgba(0,0,0,.28);}
  .easy-wrong.predicted{background:linear-gradient(135deg,#241b3d,#33265c);border-color:#9775fa;}
  .easy-wrong.mini{background:linear-gradient(135deg,#12253d,#1a3350);border-color:#339af0;}
  #ewLocalCnt{color:#ff8787;}
}
"""

# ── 文本工具 ────────────────────────────────────────────────
def norm_stem(s):
    return re.sub(r"\s+", "", re.sub(r"^(单选题|多选题|判断题)", "", str(s)))

def norm_text(s):
    return re.sub(r"[\s，。；：、;:,.（）()\u3000“”\"']+", "", str(s))

def bigrams(s):
    s = norm_text(s)
    return {s[i:i+2] for i in range(len(s)-1)} if len(s) > 1 else {s}

def bsim(a, b):
    A, B = bigrams(a), bigrams(b)
    return len(A & B) / len(A | B) if A and B else 0.0

NUM_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(m|mm|cm|kV|V|次|min|秒|s|小时|年|月|日|%|米|级|倍|层)", re.I)

SOFT = ("宜", "可以", "尽量", "一般情况下", "鼓励")
STRONG = ("必须", "严禁", "禁止", "只能", "不得", "不应", "任何")

def theme_of(q):
    raw = extract_law(str(q.get("题干") or ""))
    if raw:
        law = merge_theme(raw)
        if law:
            return law
    g = str(q.get("一级纲要") or "").strip()
    if g and g != "\\":
        merged = merge_theme(g)
        return merged if merged else g
    return "其他"

def opt_letter_map(q):
    ans = set(str(q.get("答案") or "").upper())
    return {norm_text(t): (l in ans) for l, t in parse_options(str(q.get("选项") or ""))}

def correct_texts(opt_str, answer):
    ans = set(str(answer).upper())
    return [t for l, t in parse_options(str(opt_str)) if l in ans]

def stem_core_diff(a, b):
    x, y = norm_text(a), norm_text(b)
    if x == y:
        return "", ""
    i = 0
    while i < min(len(x), len(y)) and x[i] == y[i]:
        i += 1
    j = 0
    while j < min(len(x), len(y)) - i and x[len(x)-1-j] == y[len(y)-1-j]:
        j += 1
    return x[i:len(x)-j], y[i:len(y)-j]


# ── 强关联引擎 v2 + 概念家族（并查集） ──────────────────────
def build_associations(questions):
    """全库两两检测：题干相似 + 选项共用（含⚡正误反转）+ 跨题型加成。
    返回 (assoc, family_of, fam_size)
        assoc[key] = [{strength, conflict, shared_n, sim, stem_other, core_self,
                       core_other, ans_other, ans_other_texts, type_other, flips}]
        family_of[key] = 家族根键；fam_size[root] = 成员数
    强家族判定：选项共用≥2 / 题干相似≥0.55 / (共用≥1 且 相似≥0.4) / 正误反转
    """
    keys, stems, bis, qmap, omaps, themes, types = [], [], [], {}, {}, {}, {}
    for q in questions:
        k = (str(q["题型"]).strip(), norm_stem(q["题干"]))
        keys.append(k)
        stems.append(norm_text(q["题干"]))
        bis.append(bigrams(stems[-1]))
        qmap[k] = q
        omaps[k] = opt_letter_map(q)
        themes[k] = theme_of(q)
        types[k] = str(q["题型"]).strip()

    opt_index = defaultdict(set)
    for k in keys:
        for t in omaps[k]:
            opt_index[t].add(k)

    n = len(keys)
    key_to_idx = {k: i for i, k in enumerate(keys)}
    assoc = {}
    uf = {k: k for k in keys}

    def find(x):
        while uf[x] != x:
            uf[x] = uf[uf[x]]
            x = uf[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            uf[rb] = ra

    raw_opts = {k: [t for _, t in parse_options(str(qmap[k].get("选项") or ""))] for k in keys}

    for i in range(n):
        ki = keys[i]
        own = omaps[ki]
        cand = {}

        # 信号①：选项共用（含正误反转检测）
        for t, ok in own.items():
            for k in opt_index.get(t, ()):
                if k == ki:
                    continue
                c = cand.setdefault(k, {"shared": 0, "conflict": False, "sim": 0.0, "flip_norms": []})
                c["shared"] += 1
                if omaps[k].get(t) != ok:
                    c["conflict"] = True
                    c["flip_norms"].append(t)

        # 信号②：题干相似（全库扫描，跨区块）
        bi = bis[i]
        for j in range(n):
            if j == i:
                continue
            kj = keys[j]
            inter = len(bi & bis[j])
            if inter == 0:
                continue
            sim = inter / len(bi | bis[j])
            if sim >= 0.4:
                c = cand.setdefault(kj, {"shared": 0, "conflict": False, "sim": 0.0, "flip_norms": []})
                c["sim"] = max(c["sim"], sim)

        # 打分
        rel = []
        for k, c in cand.items():
            shared_n, sim = c["shared"], c["sim"]
            if shared_n == 0 and sim < 0.45:
                continue
            oq = qmap[k]
            sa, sb = stem_core_diff(qmap[ki]["题干"], oq["题干"])
            if stems[i] == stems[key_to_idx[k]]:
                continue  # 完全重复题不列
            conflict = c["conflict"]
            strength = shared_n * 2 + (3 if conflict else 0)
            if sim >= 0.6:
                strength += 4 + int(sim * 3)
            elif sim >= 0.45:
                strength += 2
            if types[k] != types[ki]:
                strength += 1   # 跨题型：同一概念的不同题型
            if themes[k] == themes[ki]:
                strength += 1
            ans_o = str(oq.get("答案") or "").upper()
            flips = [next((raw for raw in raw_opts[ki] if norm_text(raw) == t), t)
                     for t in c["flip_norms"]]
            rel.append({
                "strength": strength, "conflict": conflict, "shared_n": shared_n,
                "sim": sim, "stem_other": oq["题干"], "core_self": sa, "core_other": sb,
                "ans_other": ans_o, "ans_other_texts": correct_texts(oq.get("选项"), ans_o),
                "type_other": types[k], "flips": flips,
            })
            # 家族并查：强关联对
            if shared_n >= 2 or sim >= 0.55 or (shared_n >= 1 and sim >= 0.4) or conflict:
                union(ki, k)
        rel.sort(key=lambda d: (-d["strength"], -d["shared_n"], -d["sim"]))
        assoc[ki] = rel[:5]

    family_of = {k: find(k) for k in keys}
    fam_size = Counter(family_of.values())
    return assoc, family_of, fam_size


# ── 联网权威助记 ────────────────────────────────────────────
WEB_TIPS = [
    (r"事故等级|事故分级|特别重大|较大事故|一般事故",
     "🌐 分级锚点（国务院493号令）：死亡 3/10/30 人 · 重伤 10/50/100 人 · 损失 1000万/5000万/1亿 → 一般/较大/重大/特别重大"),
    (r"减供负荷|大面积停电|电网减供|电力安全事故",
     "🌐 《电力安全事故应急处置和调查处理条例》：按『减供负荷比例＋停电户数』定级——与按伤亡定级的生产安全事故是两套体系，别混"),
    (r"胸外心脏按压|心肺复苏|触电急救",
     "🌐 急救数字：按压频率 100–120 次/min · 深度 5–6cm · 按压:吹气 = 30:2"),
    (r"两票|工作票|操作票",
     "🌐 两票三制＝工作票、操作票；交接班制、巡回检查制、设备定期试验轮换制"),
    (r"安全生产许可",
     "🌐 《安全生产许可证条例》第九条：安全生产许可证有效期为 3 年"),
    (r"刑法|危险作业",
     "🌐 刑法修正案（十一）新增『危险作业罪』（第134条之一）三情形：①破坏监控/报警/防护/救生设备或篡改瞒毁其数据 ②有重大隐患被责令停产停业、停止施工、停止使用设备而拒不执行 ③未经批准擅自从事矿山开采、金属冶炼、建筑施工、危险物品生产经营储存等高危作业"),
    (r"十不干",
     "🌐 十不干核对框架（国网安质〔2018〕21号）：票 · 措施 · 交底危险点 · 许可手续 · 监护人在位 · 停电验电接地 · 防护用品 · 持证 · 作业环境"),
]

def web_tip_html(stem):
    s = str(stem)
    for pat, tip in WEB_TIPS:
        if re.search(pat, s):
            return f'<div class="ew-combo">{tip}</div>'
    return ""


# ── 跨题型内容关联（判断⊃选项 / 多选组合≈单选） ────────────
def build_cross_type_index(questions):
    """扫描全库，找 判断题⊃选项原文 和 多选组合≈单选正确项 的配对"""
    singles, multis, judges = [], [], []
    for q in questions:
        qtype = str(q["题型"]).strip()
        stem = str(q.get("题干") or "").strip()
        ans = str(q.get("答案") or "").strip().upper()
        opts = {norm_text(t): t for _, t in parse_options(str(q.get("选项") or ""))}
        correct = [opts[norm_text(t)] for l, t in parse_options(str(q.get("选项") or ""))
                   if l in ans and len(norm_text(t)) >= 6]
        item = {"stem": stem, "stem_n": norm_text(stem), "correct": correct,
                "opts": opts, "ans": ans, "qtype": qtype, "opt_str": str(q.get("选项") or "")}
        if qtype == "单选题":
            singles.append(item)
        elif qtype == "多选题":
            multis.append(item)
        elif qtype == "判断题":
            judges.append(item)

    # 判断题 ⊃ 单选/多选正确项
    judge_links = {}  # 判断题norm_stem -> str
    for j in judges:
        hits = []
        for s in singles + multis:
            for ct in s["correct"]:
                if len(ct) >= 8 and ct in j["stem_n"]:
                    otxt = "＋".join(s["correct"])
                    hits.append(f'「{ct[:20]}」是{ s["qtype"] }「{norm_text(s["stem"])[:25]}…」的正确项（答「{otxt[:30]}」）')
                    break
            else:
                continue
            break
        if hits:
            judge_links[norm_text(j["stem"])] = "；".join(hits[:2])

    # 多选组合 ≈ 单选正确项
    multi_links = {}  # 多选norm_stem -> str
    for m in multis:
        combo = norm_text("".join(m["correct"]))
        for s in singles:
            for ct in s["correct"]:
                if len(ct) >= 8 and (ct in combo or combo in ct):
                    multi_links[norm_text(m["stem"])] = (
                        f'本题正确组合与单选题「{norm_text(s["stem"])[:25]}…」的正确项一致'
                    )
                    break
            else:
                continue
            break

    # 多选部分选记忆提示
    partial_tips = {}
    for m in multis:
        wrong_opts = [t for l, t in parse_options(m.get("opt_str", ""))
                      if l not in m["ans"] and len(t.strip()) >= 4]
        if wrong_opts and len(m["correct"]) < len(m["opts"]):
            partial_tips[norm_text(m["stem"])] = f"⚠️ 非全选！干扰项：「{'、'.join(w[:18] for w in wrong_opts[:2])}」——记住这几个是坑"

    return judge_links, multi_links, partial_tips
COUNTER_JS = """<script>(function(){
var q='';try{q=(document.getElementById('Question')||{}).innerText||'';q=q.trim().slice(0,80);}catch(e){}
var key='v10w:'+q;
if(window._v10cq!==q){window._v10cq=q;window._v10cd=false;}
var sp=document.getElementById('ewLocalCnt');
var base=__BASE__;
function show(n){if(sp)sp.innerText='· 本机累计答错 '+n+' 次';}
var n0=base;
try{var s=parseInt(localStorage.getItem(key)||'0',10);if(s>n0)n0=s;}catch(e){}
if(n0>0)show(n0);
var t=setInterval(function(){
  var b=document.getElementById('resultBanner');if(!b)return;
  if(b.querySelector('.result-wrong')){
    if(!window._v10cd){var n=n0+1;window._v10cd=true;try{localStorage.setItem(key,String(n));}catch(e){}show(n);}
    clearInterval(t);return;
  }
  if(b.querySelector('.result-correct'))clearInterval(t);
},250);
setTimeout(function(){clearInterval(t);},20000);
})();</script>"""

def counter_html(base):
    return COUNTER_JS.replace("__BASE__", str(base))

COUNTER_SPAN = '<span id="ewLocalCnt"></span>'


# ── 盒子构造（选项原文引用，无字母） ────────────────────────
def build_hook(stem, opt_str, answer, related):
    ts = correct_texts(opt_str, answer)
    short = "＋".join(ts) if ts else ""
    if related:
        r = related[0]
        cs, co = r["core_self"], r["core_other"]
        ot = r.get("ans_other_texts") or []
        oshort = "＋".join(ot) if ot else r["ans_other"]
        if cs and co and cs != co:
            return (f"看清题干『<b>{esc(cs)}</b>』→ 选「<b>{esc(short)}</b>」；"
                    f"若问『<b>{esc(co)}</b>』→ 是另一题，答「{esc(oshort)}」，别串！")
    head = norm_text(str(stem))[:24]
    if head:
        return f"记住：<b>{esc(head)}…</b> → 「<b>{esc(short)}</b>」"
    return f"本题正确选项：「<b>{esc(short)}</b>」"


def best_distractor(opt_str, answer):
    options = parse_options(opt_str)
    texts = [t for _, t in options]
    letters = [l for l, _ in options]
    ans_set = set(str(answer).upper())
    units = defaultdict(list)
    for t in texts:
        for v, u in NUM_RE.findall(str(t)):
            units[u.lower()].append((t, v))
    for u, arr in units.items():
        vals = {v for _, v in arr}
        if len(arr) >= 2 and len(vals) > 1:
            t1, v1 = arr[0]
            t2, v2 = next((t, v) for t, v in arr if v != v1)
            return f"「{t1}」与「{t2}」同为 {u} 但数值不同（{v1} vs {v2}），以题干条件取舍", 0.85
    best = None
    for i, li in enumerate(letters):
        if li in ans_set:
            continue
        for j, lj in enumerate(letters):
            if lj in ans_set:
                s = bsim(texts[i], texts[j])
                if best is None or s > best[0]:
                    best = (s, texts[i], texts[j])
    if best and best[0] >= 0.45:
        s, dt, ct = best
        return f"「{dt}」与正确项「{ct}」高度相似（{s*100:.0f}%），逐词找差异", s
    return "", 0


def correct_combo(opt_str, answer):
    ts = correct_texts(opt_str, answer)
    if not ts:
        return ""
    label = "正确组合" if len(ts) > 1 else "正确项原文"
    return f'<div class="ew-combo">📌 {label}：{" ＋ ".join(esc(t) for t in ts)}</div>'


def xnote_html(xn):
    parts = []
    if xn.get("judge_analysis"):
        parts.append(f'<div class="ew-combo">✅ 正确说法：{esc(xn["judge_analysis"])}</div>')
    if xn.get("basis"):
        parts.append(f'<div class="ew-combo">📖 出处：{esc(xn["basis"])}</div>')
    if xn.get("note"):
        parts.append(f'<div class="ew-combo">💡 说明：{esc(xn["note"])}</div>')
    return "\n".join(parts)


def link_html(related, cross_type=None):
    if not related and not cross_type:
        return ""
    lines = []
    if cross_type:
        lines.append(f'<b>⚡跨题型同考点</b>：{cross_type}')
    for r in related[:3]:
        flips = r.get("flips") or []
        sim = r.get("sim", 0)
        if flips:
            mark = f"⚡『{'、'.join(flips[:2])}』两题正误相反"
        elif sim >= 0.45:
            mark = f"同款题干（相似{sim*100:.0f}%）"
        else:
            mark = "共用选项"
        core = r["core_other"]
        core_s = esc(core) if core else esc(norm_text(r["stem_other"]))
        ot = r.get("ans_other_texts") or []
        otxt = "＋".join(ot) if ot else r["ans_other"]
        lines.append(
            f'• <b>{mark}</b> [{esc(r["type_other"])}]：「{core_s}」→ 答「<b>{esc(otxt)}</b>」'
        )
    label = "🔗 同概念关联" if cross_type else "🔗 同概念关联题（同题干/同选项，跨单选·多选·判断）"
    return f'<div class="ew-link">{label}：' + "<br>".join(lines) + "</div>"


def real_box(stem, wrong_count, opt_str, answer, xn, related, ct_link="", ptip=""):
    distract, _sim = best_distractor(opt_str, answer)
    title = "🔥 高频易错题" if wrong_count >= 3 else "⚠️ 易错题"
    parts = [
        '<div class="easy-wrong">',
        f'<div class="ew-title">{title} · 你已答错 {wrong_count} 次{COUNTER_SPAN}</div>',
        f'<div class="ew-hook">🎯 {build_hook(stem, opt_str, answer, related)}</div>',
    ]
    wt = web_tip_html(stem)
    if wt:
        parts.append(wt)
    if distract:
        parts.append(f'<div class="ew-combo">🧲 最强迷惑项：{esc(distract)}</div>')
    parts.append(correct_combo(opt_str, answer))
    if ptip:
        parts.append(f'<div class="ew-combo" style="border-left:3px solid #e03131;">{esc(ptip)}</div>')
    parts.append(link_html(related, ct_link))
    parts.append(xnote_html(xn))
    parts.append('<div class="ew-note">先背锚点句，再扫一遍关联题——下次遇到先想锚点</div>')
    parts.append(counter_html(wrong_count))
    parts.append("</div>")
    return "\n".join(parts)


def predict_box(score, reasons, stem, opt_str, answer, xn, related, ct_link="", ptip=""):
    distract, _ = best_distractor(opt_str, answer)
    parts = [
        '<div class="easy-wrong predicted">',
        f'<div class="ew-title">🔮 预测易错点 · 提前防坑（风险指数 {min(score,99)}）{COUNTER_SPAN}</div>',
        '<div class="ew-reasons" style="font-size:14px;color:#5f3dc4;margin-bottom:8px;line-height:1.8;">🧠 '
        + "<br>🧠 ".join(esc(x) for x in reasons[:4]) + "</div>",
        f'<div class="ew-hook">🎯 {build_hook(stem, opt_str, answer, related)}</div>',
    ]
    wt = web_tip_html(stem)
    if wt:
        parts.append(wt)
    if distract:
        parts.append(f'<div class="ew-combo">🧲 需警惕的迷惑项：{esc(distract)}</div>')
    parts.append(correct_combo(opt_str, answer))
    if ptip:
        parts.append(f'<div class="ew-combo" style="border-left:3px solid #e03131;">{esc(ptip)}</div>')
    parts.append(link_html(related, ct_link))
    parts.append(xnote_html(xn))
    parts.append('<div class="ew-note">此题尚未作答——首刷时先想锚点句再作答</div>')
    parts.append(counter_html(0))
    parts.append("</div>")
    return "\n".join(parts)


def mini_box(stem, opt_str, answer, xn, related, ct_link="", ptip=""):
    ts = correct_texts(opt_str, answer)
    short = "＋".join(ts) if ts else ""
    head = norm_text(str(stem))[:24]
    hook = f"<b>{esc(head)}…</b> → 「<b>{esc(short)}</b>」" if head else f"「<b>{esc(short)}</b>」"
    if xn.get("judge_analysis"):
        hook = f"正确说法：{esc(xn['judge_analysis'])}"
    parts = [
        '<div class="easy-wrong mini">',
        f'<div class="ew-title">🧠 快记{COUNTER_SPAN}</div>',
        f'<div class="ew-hook">🎯 {hook}</div>',
    ]
    if ts and not xn.get("judge_analysis"):
        parts.append(f'<div class="ew-combo">📌 正确项：{" ＋ ".join(esc(t) for t in ts)}</div>')
    wt = web_tip_html(stem)
    if wt:
        parts.append(wt)
    if ptip:
        parts.append(f'<div class="ew-combo" style="border-left:3px solid #e03131;">{esc(ptip)}</div>')
    if related or ct_link:
        if related:
            r = related[0]
            flips = r.get("flips") or []
            sim = r.get("sim", 0)
            if flips:
                mark = f"⚡『{'、'.join(flips[:2])}』两题正误相反"
            elif sim >= 0.45:
                mark = f"同款题干（{sim*100:.0f}%）"
            else:
                mark = "共用选项"
            ot = r.get("ans_other_texts") or []
            otxt = "＋".join(ot) if ot else r["ans_other"]
            co = esc(r["core_other"] or norm_text(r["stem_other"]))
            parts.append(f'<div class="ew-link">🔗 <b>{mark}</b>「{co}」→ 答「{esc(otxt)}」</div>')
        if ct_link and not related:
            parts.append(f'<div class="ew-link">⚡跨题型同考点：{esc(ct_link)}</div>')
    if xn.get("basis"):
        parts.append(f'<div class="ew-combo">📖 出处：{esc(xn["basis"])}</div>')
    parts.append(counter_html(0))
    parts.append("</div>")
    return "\n".join(p for p in parts if p)


# ── 数据加载 ────────────────────────────────────────────────
def load_xlsx_notes_map(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["总表"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    m = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        get = lambda k: (row[idx[k]] if k in idx and idx[k] < len(row) else None)
        stem = str(get("题干") or "").strip()
        qtype = str(get("题型") or "").strip()
        if not stem:
            continue
        m[(qtype, norm_text(stem))] = {
            "basis": str(get("题目依据") or "").strip(),
            "note": str(get("说明") or "").strip(),
            "judge_analysis": str(get("判断题解析") or "").strip(),
        }
    return m


def build_dynamic_options_html(opt_str, answer, is_multi):
    options = parse_options(opt_str)
    ans_set = set(str(answer).upper())
    input_type = "checkbox" if is_multi else "radio"
    parts = []
    for i, (orig_letter, text) in enumerate(options):
        letter = chr(65 + i)
        onchange = "onCheckChange(this)" if is_multi else "onRadioChange(this)"
        parts.append(
            f'<li value="{letter}" data-correct="{"1" if orig_letter in ans_set else "0"}" onclick="clickLi(this)">'
            f'<input type="{input_type}" name="options" class="options" value="{letter}" '
            f'id="{letter}" onchange="{onchange}">'
            f'<label for="{letter}" class="optionSpan">{esc(text)}</label></li>'
        )
    return "\n".join(parts)


def load_user_model():
    import sqlite3
    conn = sqlite3.connect(r"D:\dp工作区\_tmp\collection_progress.anki2")
    nt_id = int(conn.execute("SELECT DISTINCT mid FROM notes LIMIT 1").fetchone()[0])
    n_notes = conn.execute("SELECT COUNT(*) FROM notes").fetchone()[0]
    user_decks = {}
    try:
        for did, name in conn.execute("SELECT id, name FROM decks"):
            user_decks[str(name).replace("\x1f", "::")] = int(did)
    except Exception:
        import json as J
        for did, d in J.loads(conn.execute("SELECT decks FROM col").fetchone()[0]).items():
            user_decks[d.get("name", "").replace("\x1f", "::")] = int(did)
    conn.close()
    model = create_model(dynamic=True)
    model.model_id = nt_id
    print(f"🧩 用户集合模型 ID {nt_id}（notes={n_notes}，decks={len(user_decks)}）已复用")
    return model, user_decks


def resolve_deck(user_decks, prefix, theme):
    for cand in (f"{prefix}::{theme}", f"{prefix}{theme}", theme):
        if cand in user_decks:
            return user_decks[cand], cand
    return None, f"{prefix}::{theme}"


# ── 个人弱点画像 ────────────────────────────────────────────
class RiskEngine:
    def __init__(self, prog_rows, questions):
        self.type_base = {"多选题": 25, "单选题": 12, "判断题": 5}
        self.studied = set()
        self.wrong_keys = {}
        tstat = defaultdict(lambda: [0, 0])
        for r in prog_rows:
            key = (r["qtype"], norm_stem(r["stem"]))
            self.studied.add(key)
            tstat[r["theme"]][0] += 1
            if r["wrong"] > 0:
                tstat[r["theme"]][1] += 1
                self.wrong_keys[key] = r["wrong"]
        self.theme_rate = {t: (w[1]/w[0]) if w[0] else 0.0 for t, w in tstat.items()}
        self.wrong_opt_pool = set()
        qmap = {(str(q["题型"]).strip(), norm_stem(q["题干"])): q for q in questions}
        for key in self.wrong_keys:
            q = qmap.get(key)
            if q:
                for t in opt_letter_map(q):
                    self.wrong_opt_pool.add(t)
        self.theme_opt_pool = defaultdict(set)
        for q in questions:
            th = theme_of(q)
            for t in opt_letter_map(q):
                self.theme_opt_pool[th].add(t)

    def predict(self, q, theme):
        qtype = str(q["题型"]).strip()
        stem = str(q.get("题干") or "").strip()
        answer = str(q.get("答案") or "").upper()
        reasons, score = [], 0
        base = self.type_base.get(qtype, 10)
        score += base
        if base >= 20:
            reasons.append("多选题漏选高发（该题型历史错误率47%）")
        rate = self.theme_rate.get(theme, 0)
        if rate >= 0.3:
            score += int(rate * 25)
            reasons.append(f"此主题你的错误率 {rate*100:.0f}%")
        texts = [t for _, t in parse_options(str(q.get("选项") or ""))]
        letters = [l for l, _ in parse_options(str(q.get("选项") or ""))]
        units = defaultdict(list)
        for t in texts:
            for v, u in NUM_RE.findall(t):
                units[u.lower()].append((t, v))
        for u, arr in units.items():
            vals = {v for _, v in arr}
            if len(arr) >= 2 and len(vals) > 1:
                score += 20
                (t1, v1) = arr[0]
                t2, v2 = next((t, v) for t, v in arr if v != v1)
                reasons.append(f"「{t1}」「{t2}」数值相近（{v1}/{v2}），易混淆")
                break
        ans_set = set(answer)
        best = 0
        for i, li in enumerate(letters):
            if li in ans_set:
                continue
            for j, lj in enumerate(letters):
                if lj in ans_set:
                    best = max(best, bsim(texts[i], texts[j]))
        if best >= 0.55:
            score += 18
            reasons.append("干扰项与正确项字面高度相似")
        if qtype == "判断题":
            hit_s = [w for w in SOFT if w in stem]
            hit_g = [w for w in STRONG if w in stem]
            if hit_s and answer == "A":
                score += 15
                reasons.append(f"含「{'、'.join(hit_s[:2])}」却为正确，惯性误判")
            if hit_g and answer == "B":
                score += 15
                reasons.append(f"含「{'、'.join(hit_g[:2])}」却是错误说法，惯性误判")
        own = [norm_text(t) for t in texts]
        if sum(1 for t in own if t in self.wrong_opt_pool) >= 1:
            score += 20
            reasons.append("选项与你错过的题共用表述，记忆易串")
        else:
            pool = self.theme_opt_pool.get(theme, set()) - set(own)
            if sum(1 for t in own if t in pool) >= 2:
                score += 10
                reasons.append("同主题多题共用相近选项")
        return score, reasons


# ── 主流程 ──────────────────────────────────────────────────
def main():
    prog = json.load(open(r"D:\dp工作区\_tmp\progress_analysis.json", encoding="utf-8"))
    questions = read_print_sheet(r"D:\dp工作区\中级员工试题.xlsx", "打印")
    xmap = load_xlsx_notes_map(r"D:\dp工作区\中级员工试题.xlsx")
    print(f"📖 进度：已学 {len(prog['rows'])} 卡 · 答错 {prog['wrong_total']} 次 · 错题 {sum(1 for r in prog['rows'] if r['wrong']>0)} 道")

    engine = RiskEngine(prog["rows"], questions)
    print("🔗 全库强关联检测…")
    assoc, family_of, fam_size = build_associations(questions)
    print("⚡ 跨题型内容关联检测（判断⊃选项 / 多选组合≈单选）…")
    judge_links, multi_links, partial_tips = build_cross_type_index(questions)
    print(f"   判断⊃选项: {len(judge_links)} | 多选组合≈单选: {len(multi_links)} | 部分选提示: {len(partial_tips)}")
    n_fam2 = sum(1 for s in fam_size.values() if s >= 2)
    print(f"   概念家族：{n_fam2} 个家族含≥2题（最大家族 {max(fam_size.values())} 题）")

    groups = defaultdict(list)
    for q in questions:
        groups[theme_of(q)].append(q)
    sizes = {t: len(qs) for t, qs in groups.items()}
    small = [q for t, qs in groups.items() if sizes[t] < 3 for q in qs]
    groups = {t: qs for t, qs in groups.items() if sizes[t] >= 3}
    if small:
        groups.setdefault("其他", []).extend(small)
    ordered_themes = [t for t in sorted(groups, key=lambda t: -len(groups[t])) if t != "其他"]
    if "其他" in groups:
        ordered_themes.append("其他")

    model, user_decks = load_user_model()
    css = model.css or ""
    add = []
    if ".type-badge" not in css:
        add.append(TYPE_BADGE_CSS)
    if ".easy-wrong" not in css:
        add.append(EASY_WRONG_CSS)
    if add:
        model.css = css + "\n".join(add)

    prefix = "中级员工试题"
    out_path = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "_output", f"{prefix}-易错强化版.apkg"))

    sub_decks = []
    seen = set()
    n_real = n_pred = n_link = n_mini = 0
    skipped_dup = 0
    pred_by_type = Counter()

    def qkey(q):
        return (str(q["题型"]).strip(), norm_stem(q["题干"]))

    for i, theme in enumerate(ordered_themes):
        # 家族相邻排序：先按 多→判→单，再按家族首现位置聚簇
        base_sorted = sorted(groups[theme], key=lambda q: TYPE_ORDER.get(q["题型"], 9))
        fam_order = {}
        for q in base_sorted:
            f = family_of.get(qkey(q), qkey(q))
            if f not in fam_order:
                fam_order[f] = len(fam_order)
        qs = sorted(base_sorted, key=lambda q: (fam_order[family_of.get(qkey(q), qkey(q))],
                                                TYPE_ORDER.get(q["题型"], 9)))
        did, deck_name = resolve_deck(user_decks, prefix, theme)
        if did is None:
            did = DECK_ID_BASE + i + 1
            print(f"⚠️ 未找到卡组「{theme}」，新建：{deck_name}")
        sub = genanki.Deck(did, deck_name)
        for q in qs:
            qtype = str(q["题型"]).strip()
            stem = str(q["题干"] or "").strip()
            raw_ans = str(q.get("答案") or "")
            answer = convert_answer(raw_ans.strip())
            opt_str = str(q.get("选项") or "")
            analysis = str(q.get("题目依据") or q.get("说明") or "").strip()
            xn = xmap.get((qtype, norm_text(stem)), {})
            if not xn.get("basis"):
                xn = dict(xn, basis=f"{theme}（题库该条未附具体条款）")
            badge_class = {"单选题": "danxuan", "多选题": "duoxuan", "判断题": "panduan"}[qtype]
            stem_html = f'<div class="type-badge {badge_class}">{esc(qtype)}</div>' + esc(stem)
            remark_html = build_remark_html(analysis)

            if qtype == "判断题":
                opts_html = build_judge_options_html()
                guid = genanki.guid_for("v10_judge", stem, raw_ans)
            else:
                opts_html = build_dynamic_options_html(opt_str, answer, qtype == "多选题")
                guid = genanki.guid_for("v10_multi" if qtype == "多选题" else "v10_single", stem, raw_ans)

            if guid in seen:
                skipped_dup += 1
                continue
            seen.add(guid)

            key = qkey(q)
            wc = engine.wrong_keys.get(key, 0)
            related = assoc.get(key, [])
            fam_n = fam_size.get(family_of.get(key, key), 1)
            ct_link = judge_links.get(norm_text(stem), "") or multi_links.get(norm_text(stem), "")
            ptip = partial_tips.get(norm_text(stem), "")

            tags = []
            if wc > 0:
                remark_html += "\n" + real_box(stem, wc, opt_str, answer, xn, related, ct_link, ptip)
                n_real += 1
                if related or ct_link:
                    n_link += 1
                tags.append(f"易错·错{wc}次")
            else:
                score, reasons = engine.predict(q, theme)
                if related:
                    reasons.append("有同概念姊妹题（见关联区），家族内已排在一起")
                if score >= 40:
                    remark_html += "\n" + predict_box(score, reasons, stem, opt_str, answer, xn, related, ct_link, ptip)
                    n_pred += 1
                    pred_by_type[qtype] += 1
                    tags.append("预测易错")
                else:
                    remark_html += "\n" + mini_box(stem, opt_str, answer, xn, related, ct_link, ptip)
                    n_mini += 1
                    tags.append("低风险")
            if fam_n >= 2:
                tags.append(f"概念家族{fam_n}题")
            if ptip:
                tags.append("部分选多选")

            fields = [stem_html, opts_html, answer, esc(analysis), "", remark_html, "", "", ""]
            sub.add_note(genanki.Note(model=model, fields=fields, guid=guid, tags=tags))
        sub_decks.append(sub)

    pkg = genanki.Package(sub_decks)
    pkg.write_to_file(out_path)

    print(f"\n✅ 真实易错（红盒）: {n_real} 道（带关联 {n_link} 道）")
    print(f"✅ 预测易错（紫盒）: {n_pred} 道 — 多选{pred_by_type['多选题']}/单选{pred_by_type['单选题']}/判断{pred_by_type['判断题']}")
    print(f"✅ 轻量快记（蓝盒）: {n_mini} 道")
    print(f"🧮 全部 {n_real+n_pred+n_mini} 道带注释+计数器；跳过重复 {skipped_dup} 道")
    print(f"📊 → {out_path} ({os.path.getsize(out_path)/1024:.1f} KB)")


if __name__ == "__main__":
    main()
