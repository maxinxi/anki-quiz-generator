#!/usr/bin/env python3
"""
易错题专项练习生成器
====================
从用户最新学习进度中筛出全部答错的题，生成独立的「错题专项」卡组：
- 新 GUID 前缀（v10d_）→ 导入为全新卡片，不影响原卡组进度
- 模型 ID 对齐用户集合；卡组：易错专项::主题，按错误次数降序排题
- 注释盒全量保留（记忆锚点/迷惑项/跨题型关联/网络助记/计数器 base=真实错次）
"""

import os, re, sys, json
from collections import Counter, defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from anki_template_v10 import (
    create_model, build_judge_options_html, esc, convert_answer, parse_options,
)
import genanki

from gen_anki_enhanced import (
    norm_stem, norm_text, theme_of, build_dynamic_options_html,
    real_box, load_xlsx_notes_map, resolve_deck, load_user_model,
)
import openpyxl

DECK_ID_BASE = 1610006000  # 与其他包错开


def main():
    prog = json.load(open(r"D:\dp工作区\_tmp\progress_analysis.json", encoding="utf-8"))
    questions = read_q()
    xmap = load_xlsx_notes_map(r"D:\dp工作区\中级员工试题.xlsx")

    # 只保留答错的题
    wrong_rows = [r for r in prog["rows"] if r["wrong"] > 0]
    wrong_map = {(r["qtype"], norm_stem(r["stem"])): r for r in wrong_rows}
    qmap = {}
    for q in questions:
        qmap[(str(q["题型"]).strip(), norm_stem(q["题干"]))] = q
    print(f"📖 错题 {len(wrong_rows)} 道（来自 {prog['wrong_total']} 次答错）")

    # 用户模型 + 卡组名
    model, user_decks = load_user_model()
    css = model.css or ""
    add = []
    if ".easy-wrong" not in css:
        add.append(gen_anki_enhanced_css())
    if add:
        model.css = css + "\n".join(add)

    # 按主题分组，主题内按错误次数降序
    groups = defaultdict(list)
    for key, r in wrong_map.items():
        q = qmap.get(key)
        if q:
            q["_wc"] = r["wrong"]
            groups[theme_of(q)].append(q)

    prefix = "易错专项"
    out_path = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "_output", f"{prefix}练习.apkg"))
    sub_decks = []
    total = 0
    did_base = DECK_ID_BASE
    idx = 0
    for theme in sorted(groups, key=lambda t: -len(groups[t])):
        qs = sorted(groups[theme], key=lambda q: -q["_wc"])
        name = f"{prefix}::{theme}"
        sub = genanki.Deck(did_base + idx + 1, name)
        idx += 1
        for q in qs:
            qtype = str(q["题型"]).strip()
            stem = str(q["题干"] or "").strip()
            raw_ans = str(q.get("答案") or "")
            answer = convert_answer(raw_ans.strip())
            opt_str = str(q.get("选项") or "")
            analysis = str(q.get("题目依据") or q.get("说明") or "").strip()
            xn = xmap.get((qtype, norm_text(stem)), {})
            if not xn.get("basis"):
                xn = dict(xn, basis=f"{theme}（题库该条未附具体条款）")
            badge_class = {"单选题": "danxuan", "多选题": "duoxuan", "判断题": "panduan"}[qtype]
            stem_html = f'<div class="type-badge {badge_class}">{esc(qtype)}</div>' + esc(stem)
            remark_html = build_remark_html(analysis) if False else ""

            from gen_anki_enhanced import build_remark_html as brh
            remark_html = brh(analysis)

            if qtype == "判断题":
                opts_html = build_judge_options_html()
                guid = genanki.guid_for("v10d_judge", stem, raw_ans)
            else:
                opts_html = build_dynamic_options_html(opt_str, answer, qtype == "多选题")
                guid = genanki.guid_for("v10d_multi" if qtype == "多选题" else "v10d_single", stem, raw_ans)

            wc = q["_wc"]
            related = []
            remark_html += "\n" + real_box(stem, wc, opt_str, answer, xn, related)
            tags = ["易错专项", f"错{wc}次"]

            fields = [stem_html, opts_html, answer, esc(analysis), "", remark_html, "", "", ""]
            sub.add_note(genanki.Note(model=model, fields=fields, guid=guid, tags=tags))
            total += 1
        print(f"  {name}: {len(qs)} 题")
        sub_decks.append(sub)

    pkg = genanki.Package(sub_decks)
    pkg.write_to_file(out_path)
    print(f"\n✅ 易错专项 {total} 题 → {out_path} ({os.path.getsize(out_path)/1024:.1f} KB)")


def read_q():
    return read_print_sheet()


def read_print_sheet():
    wb = openpyxl.load_workbook(r"D:\dp工作区\中级员工试题.xlsx", read_only=True, data_only=True)
    ws = wb["打印"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        q = {h: (row[i] if i < len(row) else None) for h, i in idx.items()}
        if q.get("题干") and str(q["题干"]).strip():
            questions.append(q)
    return questions


def gen_anki_enhanced_css():
    """从 gen_anki_enhanced.py 提取 EASY_WRONG_CSS"""
    import gen_anki_enhanced
    return gen_anki_enhanced.EASY_WRONG_CSS


if __name__ == "__main__":
    main()
