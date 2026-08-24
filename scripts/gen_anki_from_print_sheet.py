#!/usr/bin/env python3
"""
中级员工试题「打印」sheet → Anki .apkg（单模板 · 按主题分组 · 动态乱序版）
=========================================================================
用户要求（2026-08-20 会话）：
    1. 整合成【一个模板】，题目左上角标注题型（单选/多选/判断）
    2. 按题干做数据分析，内容相近的单选/多选/判断集中成一个区块练习
       （按题干中"依据/根据《XXX》"提取法规/安规名分组，相近主题合并）
    3. 每个区块内练习顺序：多选题 → 判断题 → 单选题
    4. 不相近的题（零散主题）放到最后的「其他」区块
    5. 单选/多选选项动态乱序（JS 每次显示随机重排，答案按 data-correct 重算）
    6. 判断题固定 A.正确 / B.错误，不洗牌

卡组结构（单 .apkg，一个模型）：
    中级员工试题
    ├─ ::配电安规          （按题量降序，每区块内 多选→判断→单选）
    ├─ ::近3年安全事故通报
    ├─ ...
    └─ ::其他              （零散主题，放最后）

用法：
    python gen_anki_from_print_sheet.py 中级员工试题.xlsx [输出.apkg] [--sheet 打印]
"""

import os
import re
import sys
from collections import Counter, defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from anki_template_v10 import (
    create_model, build_judge_options_html, build_remark_html, esc, convert_answer,
    parse_options,
)
import genanki
import openpyxl

DECK_ID_BASE = 1610005000  # 与旧版(4000)错开，避免导入冲突

TYPE_ORDER = {"多选题": 0, "判断题": 1, "单选题": 2}  # 区块内顺序


# ── 主题提取与合并 ──────────────────────────────────────────

def extract_law(stem):
    """提取题干中的法规/安规名（依据/根据/按照/遵照《...》）"""
    m = re.search(r"(?:依据|根据|按照|遵照)[《「\u0022]?([^》」\u0022，。；、]{2,40})", stem)
    return m.group(1).strip() if m else None


def merge_theme(t):
    """合并法规名变体 → 规范主题名"""
    if not t:
        return "其他"
    t = re.sub(r"^(根据|依据|按照|遵照)", "", t)
    t = t.strip("《》「」\"'，。、；： ")
    # 近3年事故通报 合并（国内/国网系统）
    if "近3年" in t and ("安全事故" in t or "通报" in t):
        return "近3年安全事故通报"
    # 安规类：规范为 XX安规
    m = re.match(r"^(.{2,6}?安规)$", t)
    if m:
        return m.group(1)
    # 安全生产许可(证)条例 合并
    if "安全生产许可" in t and "条例" in t:
        return "安全生产许可条例"
    # 刑法 / 刑法修正案 合并
    if "刑法" in t:
        return "刑法"
    # 新《安全生产法》 合并
    if "安全生产法" in t:
        return "安全生产法"
    # 国家电网公司安全工作规定
    if "安全工作规定" in t and "国家电网" in t:
        return "国家电网公司安全工作规定"
    # 国家电网有限公司安全工作奖惩规定
    if "安全工作奖惩规定" in t:
        return "国家电网安全工作奖惩规定"
    # 隐患排查治理管理办法
    if "隐患排查治理" in t:
        return "安全隐患排查治理管理办法"
    # 十五五规划
    if "十五五" in t or "国民经济和社会发展第十五个五年规划" in t:
        return "十五五规划纲要"
    return t


def theme_of(q):
    raw = extract_law(str(q.get("题干") or ""))
    if raw:
        law = merge_theme(raw)
        if law:
            return law
    g = str(q.get("一级纲要") or "").strip()
    if g and g != "\\":
        merged = merge_theme(g)
        return merged if merged else g
    return "其他"


# ── 数据读取 ────────────────────────────────────────────────

def read_print_sheet(filepath, sheet_name="打印"):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"❌ 找不到 sheet「{sheet_name}」，现有: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    for required in ("题型", "题干", "选项", "答案"):
        if required not in idx:
            raise SystemExit(f"❌ 表头缺少列「{required}」: {headers}")
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        q = {h: (row[i] if i < len(row) else None) for h, i in idx.items()}
        if q.get("题干") and str(q["题干"]).strip():
            questions.append(q)
    return questions


# ── 卡片构建 ────────────────────────────────────────────────

def build_dynamic_options_html(opt_str, answer, is_multi):
    """动态乱序：选项保持原序存储，li 带 data-correct；JS 每次显示随机重排。"""
    options = parse_options(opt_str)
    ans_set = set(str(answer).upper())
    input_type = "checkbox" if is_multi else "radio"
    parts = []
    for i, (orig_letter, text) in enumerate(options):
        letter = chr(65 + i)
        correct_flag = "1" if orig_letter in ans_set else "0"
        onchange = "onCheckChange(this)" if is_multi else "onRadioChange(this)"
        parts.append(
            f'<li value="{letter}" data-correct="{correct_flag}" onclick="clickLi(this)">'
            f'<input type="{input_type}" name="options" class="options" value="{letter}" '
            f'id="{letter}" onchange="{onchange}">'
            f'<label for="{letter}" class="optionSpan">{esc(text)}</label></li>'
        )
    return "\n".join(parts)


TYPE_BADGE_CSS = """
/* ── 题型角标（左上角） ─────────────────────── */
.card{position:relative;}
.type-badge{position:absolute;top:8px;left:8px;z-index:9;padding:3px 12px;border-radius:14px;font-size:13px;font-weight:800;letter-spacing:1px;color:#fff;box-shadow:0 2px 6px rgba(0,0,0,.25);}
.type-badge.danxuan{background:linear-gradient(135deg,#339af0,#1971c2);}
.type-badge.duoxuan{background:linear-gradient(135deg,#f59f00,#e8590c);}
.type-badge.panduan{background:linear-gradient(135deg,#2f9e44,#2b8a3e);}
#Question{padding-top:34px;}
"""


def build_note(model, q, qtype):
    stem = str(q.get("题干") or "").strip()
    answer = str(q.get("答案") or "").strip()
    analysis = str(q.get("题目依据") or q.get("说明") or "").strip()
    baoming_html = ""
    remark_html = build_remark_html(analysis)

    badge_class = {"单选题": "danxuan", "多选题": "duoxuan", "判断题": "panduan"}.get(qtype, "danxuan")
    badge = f'<div class="type-badge {badge_class}">{esc(qtype)}</div>'
    stem_html = badge + esc(stem)

    if qtype == "判断题":
        opts_html = build_judge_options_html()
        return genanki.Note(
            model=model,
            fields=[stem_html, opts_html, convert_answer(answer), esc(analysis),
                    baoming_html, remark_html, "", "", ""],
            guid=genanki.guid_for("v10_judge", stem, answer),
        )

    is_multi = qtype == "多选题"
    opts_html = build_dynamic_options_html(str(q.get("选项") or ""), answer, is_multi)
    return genanki.Note(
        model=model,
        fields=[stem_html, opts_html, convert_answer(answer), esc(analysis), baoming_html,
                remark_html, "", "", ""],
        guid=genanki.guid_for("v10_multi" if is_multi else "v10_single", stem, answer),
    )


# ── 主流程 ──────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    filepath = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None
    sheet_name = "打印"
    if "--sheet" in sys.argv:
        sheet_name = sys.argv[sys.argv.index("--sheet") + 1]

    questions = read_print_sheet(filepath, sheet_name)
    print(f"📖 读取「{sheet_name}」sheet：{len(questions)} 题")

    # 按主题分组
    groups = defaultdict(list)
    for q in questions:
        groups[theme_of(q)].append(q)

    # 合并 <3 题的零散主题到「其他」；「其他」区块始终放最后
    sizes = {t: len(qs) for t, qs in groups.items()}
    small = [q for t, qs in groups.items() if sizes[t] < 3 for q in qs]
    groups = {t: qs for t, qs in groups.items() if sizes[t] >= 3}
    if small:
        groups["其他"] = groups.get("其他", []) + small
    ordered_themes = [t for t in sorted(groups, key=lambda t: -len(groups[t])) if t != "其他"]
    if "其他" in groups:
        ordered_themes.append("其他")

    print(f"📚 主题区块：{len(ordered_themes)} 个")
    for t in ordered_themes:
        c = len(groups[t])
        types = Counter(q["题型"] for q in groups[t])
        print(f"   {c:4d}题  多{types['多选题']} 判{types['判断题']} 单{types['单选题']}  {t}")

    # 模型：动态乱序 + 题型角标 CSS
    model = create_model(dynamic=True)
    model.css = (model.css or "") + TYPE_BADGE_CSS

    prefix = os.path.splitext(os.path.basename(filepath))[0]
    out = out or os.path.join(SCRIPT_DIR, "..", "_output", f"{prefix}.apkg")
    out = os.path.abspath(out)
    os.makedirs(os.path.dirname(out), exist_ok=True)

    sub_decks = []
    for i, theme in enumerate(ordered_themes):
        qs = sorted(groups[theme], key=lambda q: TYPE_ORDER.get(q["题型"], 9))
        sub = genanki.Deck(DECK_ID_BASE + i + 1, f"{prefix}::{theme}")
        for q in qs:
            sub.add_note(build_note(model, q, q["题型"]))
        sub_decks.append(sub)

    pkg = genanki.Package(sub_decks)
    pkg.write_to_file(out)
    total = sum(len(v) for v in groups.values())
    print(f"\n📊 总计 {total} 题 → {out} ({os.path.getsize(out)/1024:.1f} KB)")
    print("✅ 单模板 + 左上角题型角标 + 按主题分区块（区内 多选→判断→单选）+ 动态乱序")


if __name__ == "__main__":
    main()
